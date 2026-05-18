"""
Reports router — Achievement reports with CSV/Excel export + Analytics endpoint.
"""

import csv
import io
from datetime import date
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from database import get_db
from models import User, Goal, Achievement, Cycle, Checkin, UserRole
from schemas import AnalyticsResponse
from auth import get_current_user, require_role
from routers.goals import get_active_quarter

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/achievement")
def achievement_report(cycle_id: Optional[int] = None, quarter: Optional[str] = None,
                       department: Optional[str] = None, format: str = Query("json"),
                       db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    query = db.query(Achievement).options(
        joinedload(Achievement.goal).joinedload(Goal.employee)
    )
    if cycle_id:
        query = query.join(Goal).filter(Goal.cycle_id == cycle_id)
    if quarter:
        query = query.filter(Achievement.quarter == quarter)
    if department:
        if not cycle_id:
            query = query.join(Goal)
        query = query.join(User, Goal.employee_id == User.id).filter(User.department == department)

    achievements = query.all()

    rows = []
    for ach in achievements:
        goal = ach.goal
        emp = goal.employee if goal else None
        rows.append({
            "Employee Name": emp.name if emp else "",
            "Department": emp.department if emp else "",
            "Goal Title": goal.title if goal else "",
            "Thrust Area": goal.thrust_area if goal else "",
            "UoM": goal.uom_type.value if goal else "",
            "Target": goal.target_value if goal and goal.target_value else (str(goal.target_date) if goal and goal.target_date else ""),
            "Actual": ach.actual_value if ach.actual_value is not None else (str(ach.actual_date) if ach.actual_date else ""),
            "Progress Score (%)": round(ach.progress_score, 2) if ach.progress_score else 0,
            "Status": ach.status if isinstance(ach.status, str) else ach.status.value if ach.status else "",
            "Quarter": ach.quarter if isinstance(ach.quarter, str) else ach.quarter.value if ach.quarter else "",
        })

    if format == "csv":
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        content = output.getvalue()
        return StreamingResponse(
            io.BytesIO(content.encode()),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=achievement_report.csv"}
        )

    if format == "excel":
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Achievement Report"
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([row[h] for h in headers])
            # Style header
            from openpyxl.styles import Font, PatternFill
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=achievement_report.xlsx"}
        )

    return rows


@router.get("/analytics", response_model=AnalyticsResponse)
def get_analytics(cycle_id: Optional[int] = None, db: Session = Depends(get_db),
                  current_user: User = Depends(require_role(["admin"]))):
    cycle = None
    if cycle_id:
        cycle = db.query(Cycle).filter(Cycle.id == cycle_id).first()
    else:
        cycle = db.query(Cycle).filter(Cycle.is_active == True).first()

    if not cycle:
        return AnalyticsResponse(qoq_trends=[], thrust_area_dist=[], completion_heatmap=[], manager_effectiveness=[])

    # QoQ Trends by department
    departments = [d[0] for d in db.query(User.department).distinct().all() if d[0]]
    qoq_trends = []
    for dept in departments:
        dept_data = {"department": dept}
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            achs = (db.query(Achievement)
                    .join(Goal).join(User, Goal.employee_id == User.id)
                    .filter(Goal.cycle_id == cycle.id, User.department == dept, Achievement.quarter == q)
                    .all())
            scores = [a.progress_score for a in achs if a.progress_score is not None]
            dept_data[q] = round(sum(scores) / len(scores), 1) if scores else 0
        qoq_trends.append(dept_data)

    # Thrust area distribution
    goals = db.query(Goal).filter(Goal.cycle_id == cycle.id).all()
    thrust_counts = {}
    for g in goals:
        thrust_counts[g.thrust_area] = thrust_counts.get(g.thrust_area, 0) + 1
    total_goals = len(goals) or 1
    thrust_area_dist = [{"name": k, "count": v, "percentage": round(v / total_goals * 100, 1)} for k, v in thrust_counts.items()]

    # Completion heatmap
    completion_heatmap = []
    for dept in departments:
        row = {"department": dept}
        for q in ["Q1", "Q2", "Q3", "Q4"]:
            emp_ids = [u.id for u in db.query(User).filter(User.department == dept, User.is_active == True).all()]
            if not emp_ids:
                row[q] = 0
                continue
            total_with = 0
            for eid in emp_ids:
                has = db.query(Achievement).join(Goal).filter(Goal.employee_id == eid, Goal.cycle_id == cycle.id, Achievement.quarter == q).first()
                if has:
                    total_with += 1
            row[q] = round(total_with / len(emp_ids) * 100, 1)
        completion_heatmap.append(row)

    # Manager effectiveness
    managers = db.query(User).filter(User.role == UserRole.manager, User.is_active == True).all()
    manager_effectiveness = []
    for mgr in managers:
        team = db.query(User).filter(User.manager_id == mgr.id, User.is_active == True).all()
        if not team:
            continue
        total_checkins_possible = 0
        total_checkins_done = 0
        for emp in team:
            goal_ids = [g.id for g in db.query(Goal).filter(Goal.employee_id == emp.id, Goal.cycle_id == cycle.id).all()]
            if goal_ids:
                achs = db.query(Achievement).filter(Achievement.goal_id.in_(goal_ids)).all()
                total_checkins_possible += len(achs)
                for ach in achs:
                    has_checkin = db.query(Checkin).filter(Checkin.achievement_id == ach.id, Checkin.manager_id == mgr.id).first()
                    if has_checkin:
                        total_checkins_done += 1
        rate = round(total_checkins_done / total_checkins_possible * 100, 1) if total_checkins_possible else 0
        manager_effectiveness.append({"manager": mgr.name, "rate": rate})

    manager_effectiveness.sort(key=lambda x: x["rate"], reverse=True)

    return AnalyticsResponse(qoq_trends=qoq_trends, thrust_area_dist=thrust_area_dist,
                             completion_heatmap=completion_heatmap, manager_effectiveness=manager_effectiveness)
